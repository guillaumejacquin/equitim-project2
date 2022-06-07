import site
import pymongo
from pymongo import MongoClient
from openpyxl import load_workbook

# from calculs.sponsor import sponsor

#ajoute une valeud dans la collection clients qui se trouve dans la base de donnees templates
def add_value_data_base(ticker, equity, inconvénient, dividende="", siteweb="", sponsor="", yahoo=""): 
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
    
    db = cluster["templates"]
    collection = db["clients"]

    post1 = {"Ticker": ticker, "Equity" : equity, "Inconvénient": inconvénient, "Dividende": dividende, "SiteWeb": siteweb, "Sponsor": sponsor, "Yahoo": yahoo}

    collection.insert_one(post1)

# add_value_data_base("Cac50", "ca")


#montre les valeurs de la base de la collections clients de la base de données templates
def show_database():
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")

    db = cluster["templates"]
    collection = db["clients"]
    list_elements_database = []
    for x in collection.find({}, {"_id":0, "Libelle": 1, "Ticker": 1 }): 
        list_elements_database.append(x)

    return(list_elements_database)


# show_database()

def add_value():
    file_path = 'BDD_MongoDB_Projet_Brochure.xlsm'
    wb = load_workbook(file_path)
    ws = wb['Feuil1']  # or wb.active
    
    i = 2
    while True:
        ticker = "A" + str(i)
        Equity = "B" + str(i)
        inconvénient = "C" + str(i)
        dividende = "D" + str(i)
        siteweb = "E" + str(i)
        sponsor = "F" + str(i)
        yahoo = "G" + str(i)

        if (ws[ticker].value is None):
            exit (2)

        if (ws[ticker].value is None):   
            ws[ticker].value = " "
        
        add_value_data_base(ws[ticker].value, ws[Equity].value, ws[inconvénient].value, ws[dividende].value, ws[siteweb].value, ws[sponsor].value, ws[yahoo].value )
        i+=1
        
    wb.save(file_path)
# add_value()

def takeinformations(Class):
    #Recupere le sous jacent
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")

    db = cluster["templates"]
    collection = db["clients"]

    stringlongue = ""
    #remplacer par le bon element(ici ticker)



    myresults = []
    #j ajoute les tickers
    for i in Class.TSJ:
        myresults.append(collection.find({"Ticker":i}))

    compteur = 0
    for result in myresults:
            if compteur == 0:
                mot = ""
            else:
                mot = " et "
            
            test = result[0]
 
            Class.NOMSOUSJACENT = Class.NOMSOUSJACENT + mot + (test["Equity"]) #on ajoute le sous jacent + ce qu on avait avant
            Class.DIVIDENDE = Class.DIVIDENDE + mot + test["Dividende"]
            Class.SPONSOR = Class.SPONSOR + mot + test["Sponsor"]
            Class.Site = Class.Site + mot + test["SiteWeb"]
            Class.TICKER = Class.TICKER + mot + test["Ticker"]
            Class.BLOCDIVIDENDE = Class.BLOCDIVIDENDE + mot + test["Equity"] + " (" + test["Dividende"] + " ; code Bloomberg : " + test["Ticker"] +  " ;  <sponsor> : "+ test["Sponsor"] +  " ; " + test["SiteWeb"] + ")" 
            Class.Yahoo.append(test["Yahoo"])
            Class.Yahoo_value_name.append(test["Equity"])
            Class.Yahoo_value_dividende.append(test["Dividende"])

            # except Exception: #au cas ou si ca marche pas, pour éviter que ca crash
            #     Class.NOMSOUSJACENT + mot + ("ERREUR LES POTES")
            #     Class.DIVIDENDE = Class.DIVIDENDE + mot + "ERREUR"
            #     Class.SPONSOR = Class.SPONSOR + mot + "ERREUR"
            #     Class.Site = Class.Site + mot + "ERREUR"
            #     Class.TICKER = Class.TICKER + mot + "ERREUR"
                #Class.Yahoo.append(test["Yahoo"])
                
                # Class.Yahoo_value_name.append(test["Equity"])
                # Class.Yahoo_value_dividende.append("dividende inconnu")

                # Class.BLOCDIVIDENDE = "ERRORRRRRRRRRR ERRORRRRRRRRRRRRRR ERRORRRRRR"

            compteur+=1

# add_value()