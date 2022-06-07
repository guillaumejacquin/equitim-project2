import site
import pymongo
from pymongo import MongoClient
from openpyxl import load_workbook


def add_value_data_base(ticker, equity, inconvénient, dividende="", siteweb="", sponsor="", yahoo=""): 
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
    
    db = cluster["templates"]
    collection = db["clients"]

    post1 = {"Ticker": ticker, "Equity" : equity, "Inconvénient": inconvénient, "Dividende": dividende, "SiteWeb": siteweb, "Sponsor": sponsor, "Yahoo": yahoo}

    collection.insert_one(post1)
def add_value():
    file_path = 'BDD_MongoDB_Projet_Brochure.xlsm'
    wb = load_workbook(file_path)
    ws = wb['Feuil1']  # or wb.active
    
    i = 2
    while True:
        ticker = "A" + str(i)
        Equity = "B" + str(i)
        inconvénient = "C" + str(i)
        dividende = "D" + str(i)
        siteweb = "E" + str(i)
        sponsor = "F" + str(i)
        yahoo = "G" + str(i)

        if (ws[ticker].value is None):
            exit (2)

        if (ws[ticker].value is None):   
            ws[ticker].value = " "
        
        add_value_data_base(ws[ticker].value, ws[Equity].value, ws[inconvénient].value, ws[dividende].value, ws[siteweb].value, ws[sponsor].value, ws[yahoo].value )
        i+=1
        
    wb.save(file_path)
add_value()